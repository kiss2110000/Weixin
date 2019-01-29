import re
import openpyxl
from wxpy import *


bot = Bot(cache_path=True)
order = bot.groups().search(r'订单群')
province = ['北京', '天津', '上海', '重庆', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽',
            '福建', '江西', '山东', '河南', '湖北', '湖南', '广东', '海南', '四川', '贵州', '云南', '陕西',
            '甘肃', '青海', '台湾', '内蒙古', '广西壮族', '西藏', '宁夏回族', '新疆维吾尔', '香港', '澳门']
products = [""]


@bot.register(order, except_self=False)
def reply_order(msg):
    cont = msg.text
    print(cont)
    if '数量' in cont:
        # 提示所有订单的数量
        wb = openpyxl.load_workbook(filename="order_book.xlsx")
        ws = wb.active
        names = [i.value for i in tuple(ws.columns)[3] if i.value]
        print('订单共计{}条!'.format(len(names)-1))
        return '订单共计{}条!'.format(len(names)-1)
    if '订单' in cont:
        # 列出所有订单的内容
        wb = openpyxl.load_workbook(filename="order_book.xlsx")
        ws = wb.active
        rows = tuple(ws.rows)[1:]
        # print(rows)
        content = "所有订单:\n"
        for i, row in enumerate(rows):
            if row[7].value:
                style = rows[i:i+row[7].value]
                # print(style)
                prods = []
                prices = []
                numbers = []
                for each in style:
                    prods.append(each[0].value)
                    prices.append(each[1].value)
                    numbers.append(each[2].value)
                print(prices)
                print(numbers)
                p_n = [str(price)+'x'+str(number) if number != 1 else str(price) for price, number in zip(prices,numbers)]
                print(p_n)
                cells = ["、".join(prods), "、".join(p_n), row[3].value, row[4].value, row[5].value]
                info = ",".join(cells)
                # print(info)
                content += info + '\n\n'
        print(content)
        return content

    # 去除空格，中文逗号替换为英文逗号，再按逗号分离文本
    cont = cont.replace(' ', '')
    cont = cont.replace('，', ',')
    order_info = cont.split(",")

    # 买方、卖方、联系人、电话、产品名称、规格型号、数量、单价、总价（是否包含运费）、交付方式
    if len(order_info) == 5:
        order_data = []
        # 以顿号分离产品和价格
        prods = order_info[0].split("、")
        prices = order_info[1].split("、")
        if len(prods) != len(prices):
            return "错误：\n产品名数量和价格数量不相等！"

        for i in range(len(prods)):
            prod = prods[i]
            price = prices[i]
            price = price.replace('X', 'x')
            price = price.split("x")
            if len(price) == 1 and price[0].isdigit():
                price, number = int(price[0]), 1
            elif len(price) == 2 and price[0].isdigit() and price[1].isdigit():
                price, number = int(price[0]), int(price[1])
            else:
                return "错误：\n价格和数量格式不对！例:价格x数量(150x3)、价格(150)"

            if i == 0:
                prod_num = len(prods)
                name = order_info[2]
                tel = order_info[3]
                if len(tel) != 11 or tel.isdigit() is False:
                    return "错误：\n手机号码不是11位数或者不是纯数字号码！"
                addr = None
                for each in province:
                    if each in order_info[4]:
                        addr = order_info[4]
                        break
                if addr is None:
                    return "错误：\n地址中没有包含省、直辖市、自治区、特别行政区的地址!"
                date = msg.create_time.strftime('%H:%M:%S %Y-%m-%d')

                # 种类 商品 价格 数量 买家 电话 地址  日期
                order_data.append([prod, price, number, name, tel, addr, date, prod_num])
            else:
                order_data.append([prod, price, number, "", "", "", "", ""])
        print(order_data)
        # 写入数据
        wb = openpyxl.load_workbook(filename="order_book.xlsx")
        ws = wb.active
        for i, prod in enumerate(order_data):
            print(123)
            print(i)
            ws.insert_rows(2+i)
            print(prod)
            for clo in range(1, 9):
                ws.cell(row=2+i, column=clo).value = prod[clo-1]
        # 文件打开时，保存失败！无报错!
        wb.save('order_book.xlsx')

        wb = openpyxl.load_workbook(filename="order_book.xlsx")
        ws = wb.active
        # 获取姓名列

        names = [i.value for i in tuple(ws.columns)[3] if i.value]
        print(names)
        print("总共添加{}条订单!".format(len(names)-1))
        return "总共添加{}条订单!".format(len(names)-1)


embed()


