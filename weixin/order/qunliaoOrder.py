import os
import re
import openpyxl
from wxpy import *


# 订单的所有的字段(9个)
# 客户 电话 地址 商品 价格 数量 项目数 日期 账号
# client tel addr prod price num item date account

field = ['客户', '电话', '地址', '商品', '价格', '数量', '项目数', '日期', '账号']

client_index = field.index('客户')
tel_index = field.index('电话')
addr_index = field.index('地址')
prod_index = field.index('商品')
price_index = field.index('价格')
num_index = field.index('数量')
item_index = field.index('项目数')
date_index = field.index('日期')
account_index = field.index('账号')


def chickOrderExcel(filename):
    if os.path.exists(filename) is False:
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(len(field)):
            ws.cell(column=i+1, row=1, value=field[i])
            ws.cell(column=i + 1, row=1).style = 'Input'
        wb.save(filename=filename)


def contentOrder(items):
    # 获取订单行和项目数范围内的行
    order_row = items[0]
    prods, prices, numbers = [], [], []
    for each in items:
        prods.append(each[prod_index].value)
        prices.append(each[price_index].value)
        numbers.append(each[num_index].value)
    p_n = [str(price) + 'x' + str(number) if number != 1 else str(price) for price, number in zip(prices, numbers)]
    cells = ["、".join(prods), "、".join(p_n),
             order_row[client_index].value,
             order_row[tel_index].value,
             order_row[addr_index].value]

    return ",".join(cells)


def insertOrder(order_data, filename):
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.active
    # 添加空行, 分割订单
    ws.insert_rows(2)
    for i, prod in enumerate(order_data):
        ws.insert_rows(2+i)
        # 商品 价格 数量 买家 电话 地址 日期 种类 1-8
        for clo in range(len(field)):
            ws.cell(row=2+i, column=clo+1).value = prod[clo]
            ws.cell(row=2+i, column=clo+1).style = 'Note'
    # 文件打开时，保存失败！无报错!
    wb.save(filename)


def deleteOrder(filename, value=None, field_name='客户', order_index=0):
    # 列出所有订单的内容
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.active
    rows = tuple(ws.rows)[1:]  # 第一行为字段,不需要
    num = 0
    if value is None:
        for i, row in enumerate(rows):
            item_value = row[item_index].value
            if item_value:
                if order_index == num:
                    ws.delete_rows(i+2, amount=item_value + 1)
                    break
                num += 1
        wb.save(filename)
        return "删除1个订单!"
    # 获取输入字段的index
    field_index = field.index(field_name)
    num = 0
    # print('{}{}:{}'.format(field_index, field_name, value))
    # 循环每行元组，如果项目数不为空,则表明是一个订单
    for i, row in enumerate(rows):
        item_value = row[item_index].value
        if item_value:
            # 检查查询字段的值名是否与输入值相同,获取指定的订单
            field_value = row[field_index].value
            # print(field_value)
            if field_value == value:
                if order_index == num:
                    # print("{}-{}".format(i+2, item_value+1))
                    ws.delete_rows(i+2, amount=item_value+1)
                    break
                num += 1
    wb.save(filename)
    return "删除1个订单!"


def findOrder(filename, value=None, field_name='客户'):
    # 列出所有订单的内容
    ws = openpyxl.load_workbook(filename=filename).active
    rows = tuple(ws.rows)[1:]  # 第一行为字段,不需要
    if value is None:
        row = rows[0]
        item_value = row[item_index].value
        items = rows[0:item_value]
        info = contentOrder(items)
        time = items[0][date_index].value.split()[0]
        account = items[0][account_index].value
        return '末单:[{}] {}\n'.format(account, time) + info + '\n'
    # 获取输入字段的index
    field_index = field.index(field_name)
    content = ":\n"
    num = 0
    # print('{}{}:{}'.format(field_index, field_name, value))
    # 循环每行元组，如果项目数不为空,则表明是一个订单
    for i, row in enumerate(rows):
        item_value = row[item_index].value
        if item_value:
            # 检查查询字段的值名是否与输入值相同,获取指定的订单
            field_value = row[field_index].value
            # print(field_value)
            if field_value == value:
                # 获取当前行和项目数范围内的行
                items = rows[i:i+item_value]
                info = contentOrder(items)
                time = items[0][date_index].value.split()[0]
                account = items[0][account_index].value
                content += '订单:{} [{}] {}\n'.format(num, account, time) + info + '\n\n'
                num += 1
    # 末尾的换行和冒号(订单为0时)
    content = content.rstrip('\n')
    content = content.rstrip(':')
    return "找到{}个订单".format(num) + content


def numberOrder(filename):
    # 提示所有订单的数量
    ws = openpyxl.load_workbook(filename=filename).active
    clients = [client.value for client in tuple(ws.columns)[client_index] if client.value]
    return len(clients)-1


def allOrder(filename):
    # 列出所有订单的内容
    ws = openpyxl.load_workbook(filename=filename).active
    rows = tuple(ws.rows)[1:]
    content = ":\n"
    num = 0
    # 循环每行，如果项目数里有数字,说明是一条订单
    for i, row in enumerate(rows):
        item_value = row[item_index].value
        if item_value:
            # 获取当前行和项目数范围内的行
            items = rows[i:i + item_value]
            info = contentOrder(items)
            time = items[0][date_index].value.split()[0]
            account = items[0][account_index].value
            content += '订单:{} [{}] {}\n'.format(num, account, time) + info + '\n\n'
            num += 1
    # 末尾的换行和冒号(订单为0时)
    content = content.rstrip('\n')
    content = content.rstrip(':')
    return "所有订单{}个".format(num) + content


def main():
    # 获取文件名称，作为订单群的名称
    groupName = os.path.splitext(os.path.basename(os.path.abspath(__file__)))[0]
    excelFile = groupName+'.xlsx'

    # 检查文件是否存在
    chickOrderExcel(excelFile)

    # 创建机器人
    bot = Bot(cache_path=True)
    bot.messages.max_history = 10000
    # bot.messages.search('wxpy', sender=bot.self)
    order_group = bot.groups().search(r'订单群')[0]
    # order_group.send("下单工具已上线!")

    @bot.register(order_group, except_self=False)
    def reply_order(msg):
        account = msg.raw['ActualNickName']
        cont = msg.text
        print('\n{}："{}"'.format(account, cont))

        # 去除空格，中文逗号替换为英文逗号，再按逗号分离文本
        cont = cont.replace(' ', '')
        cont = cont.replace('，', ',')
        finds = cont.split(",")

        if '数量' in cont:
            # 提示所有订单的数量
            num = numberOrder(excelFile)
            print('订单共计{}条!'.format(num))
            return '订单共计{}条!'.format(num)
        elif '查找' in cont:
            # 检查命令是否为汉字
            if len(finds) == 1:
                content = findOrder(excelFile)
                print(content)
                return content
            elif len(finds) == 2 and finds[1].isalpha():
                # 查找订单
                content = findOrder(excelFile, value=finds[1])
                print(content)
                return content
            elif len(finds) == 3 and finds[1].isalpha():
                if finds[1] not in ['客户', '电话', '价格', '数量', '项目数', '账号']:
                    return '错误: 无效字段! \n例如：客户 电话 价格 数量 项目数 账号'

                content = findOrder(excelFile, value=finds[2], field_name=finds[1])
                print(content)
                return content
        elif '删除' in cont:
            print(finds)
            # 检查命令是否为汉字
            if len(finds) == 1:
                content = deleteOrder(excelFile)
                print(content)
                return content
            elif len(finds) == 2 and finds[1].isdigit():
                content = deleteOrder(excelFile,order_index=int(finds[1]))
                print(content)
                return content
            elif len(finds) == 2 and finds[1].isalpha():
                # 查找订单
                content = deleteOrder(excelFile, value=finds[1])
                print(content)
                return content
            elif len(finds) == 3 and finds[1].isalpha():
                if finds[1] not in ['客户', '电话', '价格', '数量', '项目数', '账号']:
                    return '错误: 无效字段! \n例如：客户 电话 价格 数量 项目数 账号'
                content = deleteOrder(excelFile, value=finds[2], field_name=finds[1])
                print(content)
                return content
        elif '订单' in cont:
            # 列出所有订单的内容
            content = allOrder(excelFile)
            print(content)
            return content
        elif len(finds) == 5:
            # 例如：小粉刷，150x3，冯颖，18201098888，山东德州德城区
            order_info = finds
            # 以顿号分离产品和价格
            prods = order_info[0].split("、")
            prices = order_info[1].split("、")
            if len(prods) != len(prices):
                return "错误：\n产品名数量和价格数量不相等！"
            order_data = []
            for i in range(len(prods)):
                prod = prods[i]
                price = prices[i]
                price = price.replace('X', 'x')
                price = price.split("x")
                # 如果价格和数量都为数字,说明输入正确
                if len(price) == 1 and price[0].isdigit():
                    price, number = price[0], '1'
                elif len(price) == 2 and price[0].isdigit() and price[1].isdigit():
                    price, number = price[0], price[1]
                else:
                    return "错误：\n价格和数量格式不对！例:价格x数量(150x3)、价格(150)"

                if i == 0:
                    prod_num = len(prods)
                    name = order_info[2]
                    tel = order_info[3]
                    if len(tel) != 11 or tel.isdigit() is False:
                        return "错误：\n手机号码不是11位数或者不是纯数字号码！"
                    pro = ['北京', '天津', '上海', '重庆', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽',
                           '福建', '江西', '山东', '河南', '湖北', '湖南', '广东', '海南', '四川', '贵州', '云南', '陕西',
                           '甘肃', '青海', '台湾', '内蒙古', '广西壮族', '西藏', '宁夏回族', '新疆维吾尔', '香港', '澳门']
                    addr = None
                    for each in pro:
                        if each in order_info[4]:
                            addr = order_info[4]
                            break
                    if addr is None:
                        return "错误：\n地址中没有包含省、直辖市、自治区、特别行政区的地址!"
                    date = msg.create_time.strftime('%H:%M:%S %Y-%m-%d')

                    # 客户 电话 地址 商品 价格 数量 项目数 日期 账号
                    order_data.append([name, tel, addr, prod, int(price), int(number), prod_num, date, account])
                else:
                    # 如果有多个商品，此项只写商品、数量、价格、其他为空
                    order_data.append(["", "", "", prod, int(price), int(number), "", "", ""])
            # print(order_data)

            # 写入数据, 文件打开时，保存失败！无报错!
            insertOrder(order_data, excelFile)
            # 查看订单数量
            num = numberOrder(excelFile)
            print("总共添加{}条订单!".format(num))
            return "总共添加{}条订单!".format(num)

    embed()


if __name__ == '__main__':
    main()
