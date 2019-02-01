import os
import re
import time
import openpyxl

getdata = {
    '客户': '冯颖',
    '电话': '18201098888',
    'addr': '山东德州德城区',
    'prods': [
        {
            'prod': '小粉刷',
            'price': '150',
            'number': '3',
        },
        {
            'prod': '素颜霜',
            'price': '132',
            'number': '3',
        },
    ],
    'date': '2019/02/01 16:11:35',
    'account': '青年',
}
setdata = {
    '客户': '冯颖',
    '电话': '18201098888',
    '地址': '山东德州德城区',
    '商品组': [
        {
            '商品': '小粉刷',
            '价格': '150',
            '数量': '3',
        },
        {
            '商品': '素颜霜',
            '价格': '132',
            '数量': '3',
        },
    ],
    '账号': '青年',
}



class Product(object):
    def __init__(self, name, price, number):
        self.name = name
        self.price = price
        self.number = number

    def __str__(self):
        return self.name

    def __repr__(self):
        return "商品：{}".format(self.name)


class Order(object):
    def __init__(self, client, tel, addr ,products, date, account):
        self.client = client
        self.tel = tel
        self.addr = addr
        self.products = products
        self.date = date
        self.account = account

    # def __str__(self):
    #     return '< 订单:{}  提交账号:{} >'.format(self.client, self.account)

    def get_info(self):
        prod = []
        pr_nb = []
        for product in self.products:
            prod.append(product.name)
            pr_nb.append(str(product.price) + 'x' + str(product.number))
        prod = '、'.join(prod)
        pr_nb = "、".join(pr_nb)
        return prod + ',' + pr_nb + ',' + self.client + ',' + self.tel + ',' + self.addr

    def save_excel(self, file):
        pass


class OrderStore(object):
    __fields = ('客户', '电话', '地址', '商品', '价格', '数量', '项目数', '日期', '账号')
    styles = {'客户': 'Note',
              '电话': 'Note',
              '地址': 'Note',
              '商品': 'Input',
              '价格': 'Input',
              '数量': 'Input',
              '项目数': 'Note',
              '日期': 'Note',
              '账号': 'Note',
              }

    def __init__(self, filename):
        self.file = filename
        if os.path.exists(filename) is False:
            wb = openpyxl.Workbook()
            ws = wb.active
            for i in range(len(OrderStore.__fields)):
                ws.cell(column=i + 1, row=1, value=OrderStore.__fields[i])
                ws.cell(column=i + 1, row=1).style = 'Check Cell'
            wb.save(filename=filename)

    def insert(self, data):
        prods = data['商品组']
        # print(prods)
        data['项目数'] = len(prods)
        data['日期'] = time.strftime('%H:%M:%S %Y-%m-%d', time.localtime())

        wb = openpyxl.load_workbook(filename=self.file)
        ws = wb.active
        fields = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column+1)]
        # print(fields)

        rows = []
        for i in range(data['项目数']):
            dict_temp = []
            for field in fields:
                if field in data.keys():
                    if i == 0:
                        dict_temp.append(data[field])
                    else:
                        dict_temp.append("")
                else:
                    if field == '价格':
                        dict_temp.append(float(prods[i][field]))
                    elif field == '数量':
                        dict_temp.append(int(prods[i][field]))
                    else:
                        dict_temp.append(prods[i][field])
            rows.append(dict_temp)
        # print(rows)
        # 添加空行, 分割订单
        ws.insert_rows(2)
        for i, prod in enumerate(rows):
            ws.insert_rows(2 + i)
            # 商品 价格 数量 买家 电话 地址 日期 种类 1-8
            for o in range(len(fields)):
                ws.cell(row=2 + i, column=o + 1).value = prod[o]
                ws.cell(row=2 + i, column=o + 1).style = 'Note'
                ws.cell(row=2 + i, column=o + 1).style = OrderStore.styles[fields[o]]
        # 文件打开时，保存失败！无报错!
        wb.save(self.file)

OrderStore('qunliaoOrder.xlsx').insert(setdata)
# patterns = {
#     'insert': re.compile(r'(.+),(\d{1,3}x\d{1,3}),(\w{1,8}),(\d{11}),(.*)'),
# }
#
#
# text = '小粉刷、鞋子、大师傅，150x3、45、89x4，冯颖，18201098888，山东德州德城区'
# text = text.replace(' ', '')
# text = text.replace('，', ',')
# se = re.search(r'(?P<prods>.+(、.+)*),(?P<prnb>\d{1,3}(x\d{1,3})?(、\d{1,3}(x\d{1,3})?)*),(?P<client>\w{1,8}),(?P<tel>\d{11}),(?P<addr>.*)', text)
# if se:
#     print(se.groupdict())
# finds = text.split(",")
#
# order_info = finds
# # 以顿号分离产品和价格
# prods = order_info[0].split("、")
# temps = order_info[1].split("、")
# for i in range(len(prods)):
#     price, number = None, None
#     prod = prods[i]
#     pr_nb = temps[i].replace('X', 'x').split("x")
#
#     if len(pr_nb) == 1 and pr_nb[0].isdigit():
#         price, number = pr_nb[0], '1'
#     elif len(pr_nb) == 2 and pr_nb[0].isdigit() and pr_nb[1].isdigit():
#         price, number = pr_nb[0], pr_nb[1]
#     # else:
#     #     return "错误：\n价格和数量格式不对！例:价格x数量(150x3)、价格(150)"
#     prods[i] = Product(prod, int(price), int(number))
# client = order_info[2]
# tel = order_info[3]
# # if len(tel) != 11 or tel.isdigit() is False:
# #     return "错误：\n手机号码不是11位数或者不是纯数字号码！"
# pro = ['北京', '天津', '上海', '重庆', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽',
#        '福建', '江西', '山东', '河南', '湖北', '湖南', '广东', '海南', '四川', '贵州', '云南', '陕西',
#        '甘肃', '青海', '台湾', '内蒙古', '广西壮族', '西藏', '宁夏回族', '新疆维吾尔', '香港', '澳门']
# addr = None
# for each in pro:
#     if each in order_info[4]:
#         addr = order_info[4]
#         break
# # if addr is None:
# #     # return "错误：\n地址中没有包含省、直辖市、自治区、特别行政区的地址!"
# date = '246554'
# account = 'adf[p'
#
# order = Order(client, tel, addr, prods, date, account)
# # print(order.get_info())
